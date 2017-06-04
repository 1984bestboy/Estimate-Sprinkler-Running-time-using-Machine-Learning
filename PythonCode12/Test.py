
# Reference - https://stackoverflow.com/questions/26946337/reading-serial-data-from-arduino-project-pyserial
# Reference - https://stackoverflow.com/questions/13381384/modify-an-existing-excel-file-using-openpyxl-in-python?rq=1

import serial
import sys, time
from threading import Timer
from sklearn.ensemble import RandomForestRegressor


def main_code():
    port = "/dev/cu.usbmodem1421"
    baudrate = 9600
    i = 0

    total_mositure_level = 0.0
    total_humidity = 0.0
    total_temperature = 0.0

    if len(sys.argv) == 3:
        ser = serial.Serial(sys.argv[1], sys.argv[2])
    else:
        print "# Please specify a port and a baudrate"
        print "# using hard coded defaults " + port + " " + str(baudrate)
        ser = serial.Serial(port, baudrate)

    # New Code
    from openpyxl import load_workbook


    #Open an xlsx for reading
    wb = load_workbook("/Users/balaji/Documents/Github/IOT/Weather/src/newFile.xlsx")

    # Get the sheet name Current Weather
    ws = wb.get_sheet_by_name("Current Weather")

    wb.save("/Users/balaji/Documents/Github/IOT/Weather/src/newFile.xlsx")

    last_row=ws.max_row

    print "Current Value : " + ser.readline()

    string_line = ser.readline()
    list_values = string_line.split(',')

    j = 0

    for item in list_values:

        j = j + 1

        if (j == 1):
            print "Humdity :" + item
            total_humidity = total_humidity + float(item)

        if (j == 2):
            print "Temperature : " + item
            total_temperature = total_temperature + float(item)
        if (j == 3):
            print "Moisture : " + item
            total_mositure_level = total_mositure_level + float(item)

    ws.cell(row=last_row,column=13).value = (total_humidity / 10)
    ws.cell(row=last_row,column=14).value = (total_temperature / 10)
    ws.cell(row=last_row,column=15).value = (total_mositure_level / 10)

    wb.save("/Users/balaji/Documents/Github/IOT/Weather/src/newFile.xlsx")
    #t = Timer(13, main_code)
    #t.start()

def machine_learning():
    print "Machine Learning code started"

    import pandas as pd
    from openpyxl import load_workbook

    forecasted_temperature = 0.0000
    forecasted_humidity = 0.0000
    forecasted_wind_speed = 0.0000
    data = pd.read_excel('/Users/balaji/Documents/Github/IOT/Weather/src/newFile.xlsx', sheetname='Current Weather')

    print(data)

    data.head()
    X = data[['Forecasted_Temperature', 'Forecasted_Humidity', 'Forecasted_Wind_Speed']]

    X.head

    y = data['Current Moisture']

    y.head


    rfr = RandomForestRegressor(n_estimators=10)
    rfr.fit(X, y, sample_weight=None)

    wb = load_workbook("/Users/balaji/Documents/Github/IOT/Weather/src/newFile.xlsx")

    ws = wb.get_sheet_by_name("Current Weather")

    last_row = ws.max_row - 1

    forecasted_temperature = ws.cell(row=last_row + 1, column=8).value
    forecasted_humidity = ws.cell(row=last_row + 1, column=9).value
    forecasted_wind_speed = ws.cell(row=last_row + 1, column=12).value

    X_test = [forecasted_temperature, forecasted_humidity, forecasted_wind_speed]

    y_pred = rfr.predict(X_test)

    print(y_pred)
    write_arduino_tmp(y_pred)

    type(y_pred)

    X_test.append(y_pred)

    print(X_test)

    ws.cell(row=last_row + 1, column=16).value = (y_pred[0])
    wb.save("/Users/balaji/Documents/Github/IOT/Weather/src/newFile.xlsx")


def write_to_arduino(input_value):
    send_value=""
    print "Reading and writing to Arduino Started"
    arduino = serial.Serial('/dev/cu.usbmodem1421', 9600, timeout=.1)
    time.sleep(2)
    print "Input Value"

    send_value = input_value
    arduino.write("python")
    print send_value
    #while True:
    #time.sleep(2)
    data = arduino.readline()
    print "The line has been read " + data +" value"
    if data:
        #print data.rstrip('\n')  # strip out the new lines for now
        print "The data is: "
        print  (data)
        #print data
        # (better to do .read() in the long run for this reason

def write_arduino_tmp(input):
    import serial, time
    arduino = serial.Serial('/dev/cu.usbmodem1421', 9600, timeout=.1)
    time.sleep(1)  # give the connection a second to settle
    input_string = ""
    input_value = 1254
    input_string = str(input)
    l=0
    while (l<1):
        l=l+1
        arduino.write(input_string)
        time.sleep(1)
        data = arduino.readline()

        print "data is :" + data
        # if data:
        # print data #strip out the new lines for now
        # (better to do .read() in the long run for this reason

while 1:
    time.sleep(10)
    main_code()
    time.sleep(10)
    machine_learning()
    time.sleep(71)
# time.sleep(10)
# write_arduino_tmp(




