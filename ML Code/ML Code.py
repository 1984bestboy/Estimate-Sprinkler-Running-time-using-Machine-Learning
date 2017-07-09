
import pandas as pd
from openpyxl import load_workbook



forecasted_temperature = 0.0000
forecasted_humidity = 0.0000
forecasted_wind_speed=0.0000
data = pd.read_excel('/Users/balaji/Documents/Github/IOT/Weather/src/newFile.xlsx', sheetname='Current Weather')

data.head()
X=data[['Forecasted_Temperature','Forecasted_Humidity','Forecasted_Wind_Speed']]

X.head

y=data['Current Moisture']

y.head


#Defining the model
from sklearn.ensemble import RandomForestRegressor
rfr = RandomForestRegressor(n_estimators=10)
rfr.fit(X,y, sample_weight=None)

wb = load_workbook("/Users/balaji/Documents/Github/IOT/Weather/src/newFile.xlsx")

ws = wb.get_sheet_by_name("Current Weather")

last_row=ws.max_row-1

forecasted_temperature = ws.cell(row=last_row+1,column=8).value
forecasted_humidity = ws.cell(row=last_row+1,column=9).value
forecasted_wind_speed=ws.cell(row=last_row+1,column=12).value

X_test = [forecasted_temperature,forecasted_humidity,forecasted_wind_speed]

y_pred=rfr.predict(X_test)

print("The predicted value is " + y_pred)


X_test.append(y_pred)

#Writing back to the excel sheet
ws.cell(row=last_row+1,column=16).value = (y_pred[0])
wb.save("/Users/balaji/Documents/Github/IOT/Weather/src/newFile.xlsx")
